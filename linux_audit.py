#!/usr/bin/env python3
"""
linux_audit.py - SSH fleet inventory & security-hardening baseline collector.

Reads a list of hosts from a text file, connects over SSH (reusing your existing
keys / ssh-agent / ~/.ssh/config), escalates privileges with sudo, runs ONE
remote collection pass per host, and writes a formatted multi-sheet .xlsx report.

You must only run this against hosts you are authorised to access.

Requires locally: python3, the system `ssh` client, and `openpyxl`
    pip install openpyxl

Typical use:
    python3 linux_audit.py -H hosts.txt -o audit.xlsx -u admin --sudo
    python3 linux_audit.py -H hosts.txt --sudo --ask-sudo-pass --deep

hosts.txt format (one per line; blank lines and #comments ignored):
    web01.example.com
    admin@db01.example.com
    10.0.0.5:2222            # host:port
    bastion.example.com:22   # inline comments ok
"""

import argparse
import base64
import os
import concurrent.futures as cf
import getpass
import re
import shlex
import subprocess
import sys
from datetime import datetime

BANNER = """
        ✠ THE CHURCH OF THE ETERNAL CLUSTER ✠

     █████╗  █████╗     █████╗  █████╗  █████╗
    ██╔══██╗██╔══██╗   ██╔══██╗██╔══██╗██╔══██╗
    ╚██████║╚██████║   ╚██████║╚██████║╚██████║
     ╚═══██║ ╚═══██║    ╚═══██║ ╚═══██║ ╚═══██║
     █████╔╝ █████╔╝██╗ █████╔╝ █████╔╝ █████╔╝
     ╚════╝  ╚════╝ ╚═╝ ╚════╝  ╚════╝  ╚════╝

        F I V E   N I N E S ,   A M E N .

   ┌─────────────────────────────────────────┐
   │  $ kubectl get salvation                │
   │  NAME        READY   STATUS    RESTARTS │
   │  thy-soul    1/1     Running   0        │
   └─────────────────────────────────────────┘

     "And the Scheduler saw the pod, and it was Good."
                            — Book of Deployments 1:1

          ⛪  https://HA-llelujah.dev  ⛪
"""


def print_banner():
    try:
        print(BANNER, file=sys.stderr)
    except UnicodeEncodeError:
        print("THE CHURCH OF THE ETERNAL CLUSTER — 99.999 — FIVE NINES, AMEN.", file=sys.stderr)
        print("https://HA-llelujah.dev", file=sys.stderr)

# ----------------------------------------------------------------------------
# Remote collector. Runs on each target. Pure POSIX-ish bash, defensive:
# every probe is guarded so a missing tool never aborts the run. Output is
# split into "@@SECTION name" blocks that the Python side parses.
# NOTE: kept as a plain string (NOT an f-string) so $ and () are left intact.
# ----------------------------------------------------------------------------
COLLECTOR = r'''
export LC_ALL=C
emit(){ printf '@@SECTION %s\n' "$1"; }
have(){ command -v "$1" >/dev/null 2>&1; }

emit hostname
( hostname -f 2>/dev/null || hostname 2>/dev/null || cat /proc/sys/kernel/hostname 2>/dev/null )

emit whoami
id 2>/dev/null

emit os_release
cat /etc/os-release 2>/dev/null || cat /etc/redhat-release 2>/dev/null

emit kernel
uname -sr 2>/dev/null
uname -m 2>/dev/null

emit uptime
uptime -p 2>/dev/null
cat /proc/uptime 2>/dev/null
cat /proc/loadavg 2>/dev/null

emit ipaddr
if have ip; then ip -o addr show scope global 2>/dev/null | awk '{print $2, $3, $4}';
else hostname -I 2>/dev/null; fi

emit disk
df -hPT 2>/dev/null | grep -viE '^(Filesystem|tmpfs|devtmpfs|udev|overlay|squashfs)'

emit mem
if have free; then free -m 2>/dev/null;
else grep -E 'MemTotal|MemAvailable|SwapTotal' /proc/meminfo 2>/dev/null; fi

emit listening
if have ss; then ss -tulpnH 2>/dev/null;
elif have netstat; then netstat -tulpn 2>/dev/null | grep -iE 'LISTEN|udp'; fi

emit services
if have systemctl; then systemctl list-units --type=service --state=running --no-legend --no-pager 2>/dev/null | awk '{print $1}';
else service --status-all 2>/dev/null; fi

emit software
for b in sshd nginx httpd apache2 mysqld mariadbd mysql psql postgres postgresql docker containerd kubelet php python3 perl openssl redis-server mongod named vsftpd postfix dovecot haproxy java node smbd; do
  if have "$b"; then
    v="$("$b" --version 2>&1 | head -1)"
    [ -z "$v" ] && v="$("$b" -v 2>&1 | head -1)"
    [ -z "$v" ] && v="$("$b" -V 2>&1 | head -1)"
    printf '%s\t%s\n' "$b" "$v"
  fi
done
if have ssh; then printf 'openssh\t%s\n' "$(ssh -V 2>&1 | head -1)"; fi

emit users
awk -F: '{print $1":"$3":"$4":"$7}' /etc/passwd 2>/dev/null

emit uid0
awk -F: '($3==0){print $1}' /etc/passwd 2>/dev/null

emit sudoers
getent group sudo wheel admin 2>/dev/null

emit passwd_status
if have passwd; then
  awk -F: '{print $1}' /etc/passwd 2>/dev/null | while read -r u; do passwd -S "$u" 2>/dev/null; done
fi

emit ssh_effective
if have sshd; then sshd -T 2>/dev/null | grep -iE '^(permitrootlogin|passwordauthentication|permitemptypasswords|x11forwarding|maxauthtries|challengeresponseauthentication|kbdinteractiveauthentication|pubkeyauthentication|clientaliveinterval|logingracetime|allowtcpforwarding|permituserenvironment) '; fi

emit ssh_file
grep -iE '^[[:space:]]*(permitrootlogin|passwordauthentication|permitemptypasswords|x11forwarding|maxauthtries|port|protocol)' /etc/ssh/sshd_config 2>/dev/null

emit firewall
if have ufw; then printf 'ufw:%s\n' "$(ufw status 2>/dev/null | head -1)"; fi
if have firewall-cmd; then printf 'firewalld:%s\n' "$(firewall-cmd --state 2>/dev/null)"; fi
if have nft; then printf 'nftables_lines:%s\n' "$(nft list ruleset 2>/dev/null | grep -c .)"; fi
if have iptables; then printf 'iptables_rules:%s\n' "$(iptables -S 2>/dev/null | grep -c .)"; fi

emit selinux
if have getenforce; then printf 'selinux:%s\n' "$(getenforce 2>/dev/null)";
elif [ -f /sys/fs/selinux/enforce ]; then printf 'selinux:present\n';
else printf 'selinux:absent\n'; fi
if have aa-status; then
  if aa-status --enabled 2>/dev/null; then printf 'apparmor:enabled\n'; else printf 'apparmor:disabled\n'; fi
elif [ -d /sys/kernel/security/apparmor ]; then printf 'apparmor:present\n';
else printf 'apparmor:absent\n'; fi

emit updates
if have apt-get; then
  printf 'mgr:apt\n'
  printf 'pending:%s\n' "$(apt-get -s -o Debug::NoLocking=true upgrade 2>/dev/null | grep -c '^Inst')"
  printf 'security:%s\n' "$(apt-get -s -o Debug::NoLocking=true upgrade 2>/dev/null | grep '^Inst' | grep -ci 'security\|-security')"
elif have dnf; then
  printf 'mgr:dnf\n'
  printf 'pending:%s\n' "$(dnf -q check-update 2>/dev/null | grep -cE '^[[:alnum:]]')"
  printf 'security:%s\n' "$(dnf -q updateinfo list security 2>/dev/null | grep -cE '^[[:alnum:]]')"
elif have yum; then
  printf 'mgr:yum\n'
  printf 'pending:%s\n' "$(yum -q check-update 2>/dev/null | grep -cE '^[[:alnum:]]')"
fi

emit reboot_required
if [ -f /var/run/reboot-required ] || [ -f /run/reboot-required ]; then printf 'yes\n';
elif have needs-restarting; then
  if needs-restarting -r >/dev/null 2>&1; then printf 'no\n'; else printf 'yes\n'; fi
else printf 'unknown\n'; fi

emit sysctl
for k in net.ipv4.ip_forward kernel.randomize_va_space net.ipv4.tcp_syncookies net.ipv4.conf.all.rp_filter kernel.kptr_restrict kernel.dmesg_restrict net.ipv4.conf.all.accept_redirects; do
  printf '%s=%s\n' "$k" "$(sysctl -n "$k" 2>/dev/null)"
done

emit timesync
if have timedatectl; then timedatectl show -p NTP -p NTPSynchronized 2>/dev/null; fi

emit autoupdate
if [ -f /etc/apt/apt.conf.d/20auto-upgrades ]; then printf 'unattended-upgrades:configured\n';
elif have dnf-automatic; then printf 'dnf-automatic:present\n';
elif systemctl is-enabled dnf-automatic.timer >/dev/null 2>&1; then printf 'dnf-automatic:enabled\n';
else printf 'auto-updates:none-detected\n'; fi

emit pkgcount
if have dpkg-query; then dpkg-query -f '.\n' -W 2>/dev/null | wc -l;
elif have rpm; then rpm -qa 2>/dev/null | wc -l; fi

emit lastlogins
if have last; then last -n 6 -w 2>/dev/null | head -6; fi

emit logindefs
grep -E '^[[:space:]]*(PASS_MAX_DAYS|PASS_MIN_DAYS|PASS_WARN_AGE|UMASK|ENCRYPT_METHOD)' /etc/login.defs 2>/dev/null

emit insecure_pkgs
for p in telnet telnetd telnet-server rsh-client rsh-server rsh-redone-server talk talkd nis ypserv xinetd tftp tftp-server tftpd; do
  if have dpkg-query; then
    dpkg-query -W -f='${db:Status-Abbrev}' "$p" 2>/dev/null | grep -q '^ii' && printf '%s\n' "$p"
  elif have rpm; then
    rpm -q "$p" >/dev/null 2>&1 && printf '%s\n' "$p"
  fi
done

emit file_perms
for f in /etc/ssh/sshd_config /etc/passwd /etc/shadow /etc/group /etc/gshadow /etc/crontab; do
  [ -e "$f" ] && stat -c '%n %a %U %G' "$f" 2>/dev/null
done

emit packages
if [ "$PKGS" = "1" ]; then
  if have dpkg-query; then dpkg-query -W -f='${Package}\t${Version}\n' 2>/dev/null | sort;
  elif have rpm; then rpm -qa --qf '%{NAME}\t%{VERSION}-%{RELEASE}\n' 2>/dev/null | sort; fi
fi

if [ "$DEEP" = "1" ]; then
  emit suid
  timeout 90 find / -xdev \( -perm -4000 -o -perm -2000 \) -type f 2>/dev/null
  emit worldwritable
  timeout 90 find / -xdev -type f -perm -0002 2>/dev/null | head -300
fi

printf '@@END\n'
'''

RED, ORANGE, YELLOW, GREEN, GREY = "C00000", "ED7D31", "FFC000", "70AD47", "D9D9D9"

BUILD = "2026-07-09.formula-safe"
__version__ = "1.0.0"


def _guard_formula(cell):
    """Force cell to string type if its text starts with a spreadsheet
    formula/injection trigger (=, +, -, @). Prevents Excel 'repair' warnings
    and neutralises formula injection from untrusted remote host data."""
    v = cell.value
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        cell.data_type = "s"


# ----------------------------------------------------------------------------
# Host list parsing
# ----------------------------------------------------------------------------
def parse_hosts(path):
    hosts = []
    with open(path) as fh:
        for raw in fh:
            line = raw.split("#", 1)[0].strip()
            if not line:
                continue
            user = None
            port = None
            if "@" in line:
                user, line = line.split("@", 1)
            # host:port (avoid tripping on IPv6 by requiring a plain :digits tail)
            m = re.match(r"^(.*):(\d+)$", line)
            if m and m.group(1).count(":") == 0:
                line, port = m.group(1), m.group(2)
            hosts.append({"target": line.strip(), "user": user, "port": port})
    return hosts


# ----------------------------------------------------------------------------
# SSH execution
# ----------------------------------------------------------------------------
def build_ssh_base(host, args, ssh_pass):
    cmd = []
    if ssh_pass is not None:
        # -e reads the SSH password from the SSHPASS env var (never argv)
        cmd += ["sshpass", "-e"]
    cmd += ["ssh"]
    cmd += ["-o", f"ConnectTimeout={args.connect_timeout}"]
    cmd += ["-o", f"StrictHostKeyChecking={args.host_key_checking}"]
    if ssh_pass is None:
        # BatchMode fails fast on key auth; must be OFF for password auth
        cmd += ["-o", "BatchMode=yes"]
    else:
        cmd += ["-o", "PubkeyAuthentication=no",
                "-o", "PreferredAuthentications=password,keyboard-interactive"]
    if args.identity:
        cmd += ["-i", args.identity]
    port = host["port"] or args.port
    if port:
        cmd += ["-p", str(port)]
    for opt in args.ssh_opt or []:
        cmd += ["-o", opt]
    user = host["user"] or args.user
    target = f"{user}@{host['target']}" if user else host["target"]
    cmd.append(target)
    return cmd


def run_ssh(cmd, input_bytes, timeout, env=None):
    return subprocess.run(
        cmd, input=input_bytes, stdout=subprocess.PIPE,
        stderr=subprocess.PIPE, timeout=timeout, env=env,
    )


def collect_host(host, args, script_bytes, sudo_pass, ssh_pass):
    """Return (raw_output_str, error_str_or_None)."""
    base = build_ssh_base(host, args, ssh_pass)
    env = {**os.environ, "SSHPASS": ssh_pass} if ssh_pass is not None else None
    try:
        if args.escalate == "none":
            r = run_ssh(base + ["bash -s"], script_bytes, args.cmd_timeout, env=env)
            out = r.stdout.decode(errors="replace")
            if "@@SECTION" not in out:
                return out, (r.stderr.decode(errors="replace").strip() or "no data returned")
            return out, None

        # escalate == sudo
        if sudo_pass is None:
            # passwordless sudo
            r = run_ssh(base + ["sudo -n bash -s"], script_bytes, args.cmd_timeout, env=env)
            out = r.stdout.decode(errors="replace")
            if "@@SECTION" in out:
                return out, None
            err = r.stderr.decode(errors="replace").strip()
            if "password is required" in err or "a password is required" in err.lower():
                return "", "passwordless sudo failed - re-run with --ask-sudo-pass"
            return out, (err or "no data returned")

        # password sudo: stage script to a temp file, then run under sudo -S
        token = base64.urlsafe_b64encode(os.urandom(9)).decode().rstrip("=")
        remote_tmp = f"/tmp/.audit_{token}.sh"
        stage = run_ssh(base + [f"umask 077; cat > {remote_tmp}"], script_bytes,
                        args.cmd_timeout, env=env)
        if stage.returncode != 0:
            return "", "failed to stage script: " + stage.stderr.decode(errors="replace").strip()
        run_cmd = f"sudo -S -p '' bash {remote_tmp}; rc=$?; rm -f {remote_tmp}; exit $rc"
        r = run_ssh(base + [run_cmd], (sudo_pass + "\n").encode(), args.cmd_timeout, env=env)
        out = r.stdout.decode(errors="replace")
        if "@@SECTION" in out:
            return out, None
        return out, (r.stderr.decode(errors="replace").strip() or "no data returned")

    except subprocess.TimeoutExpired:
        return "", f"timed out after {args.cmd_timeout}s"
    except Exception as exc:  # noqa: BLE001
        return "", f"{type(exc).__name__}: {exc}"


# ----------------------------------------------------------------------------
# Parsing of collector output
# ----------------------------------------------------------------------------
def split_sections(raw):
    sections = {}
    current = None
    for line in raw.splitlines():
        if line.startswith("@@SECTION "):
            current = line[len("@@SECTION "):].strip()
            sections[current] = []
        elif line.strip() == "@@END":
            break
        elif current is not None:
            sections[current].append(line)
    return sections


def _first(lines):
    for l in lines:
        if l.strip():
            return l.strip()
    return ""


def parse_os_release(lines):
    kv = {}
    for l in lines:
        if "=" in l:
            k, v = l.split("=", 1)
            kv[k] = v.strip().strip('"')
    return kv.get("PRETTY_NAME") or (kv.get("NAME", "") + " " + kv.get("VERSION", "")).strip() or _first(lines)


def parse_uptime(lines):
    pretty = ""
    for l in lines:
        if l.startswith("up ") or l.strip().startswith("up "):
            pretty = l.strip()
    if not pretty:
        # derive from /proc/uptime seconds
        for l in lines:
            parts = l.split()
            if parts and re.match(r"^\d+\.\d+$", parts[0]):
                secs = int(float(parts[0]))
                d, rem = divmod(secs, 86400)
                h, rem = divmod(rem, 3600)
                m, _ = divmod(rem, 60)
                pretty = f"up {d}d {h}h {m}m"
                break
    load = ""
    for l in lines:
        p = l.split()
        if len(p) >= 3 and all(re.match(r"^\d+\.\d+$", x) for x in p[:3]):
            load = " ".join(p[:3])
    return pretty or "unknown", load


def parse_ipaddr(lines):
    ip_re = re.compile(r"^(?:\d{1,3}(?:\.\d{1,3}){3}|[0-9a-fA-F:]+:[0-9a-fA-F:]*)(?:/\d+)?$")
    ips = []
    for l in lines:
        for tok in l.split():
            if tok in ("inet", "inet6") or tok.endswith(":"):
                continue
            if ip_re.match(tok):
                ips.append(tok.split("/")[0])
    # keep order, dedupe, drop loopback
    seen, out = set(), []
    for ip in ips:
        if ip in seen or ip.startswith("127.") or ip == "::1":
            continue
        seen.add(ip)
        out.append(ip)
    return out


def parse_disk(lines):
    rows = []
    for l in lines:
        f = l.split(None, 6)
        if len(f) >= 7 and f[5].endswith("%"):
            rows.append({
                "filesystem": f[0], "type": f[1], "size": f[2], "used": f[3],
                "avail": f[4], "use_pct": f[5], "mount": f[6],
            })
    return rows


def parse_listening(lines):
    rows = []
    for l in lines:
        if not l.strip():
            continue
        parts = l.split()
        proto = parts[0] if parts else ""
        local = ""
        # ss -H columns: Netid State Recv-Q Send-Q Local Peer Process
        m = re.findall(r"(\S+:\d+)\s", l + " ")
        if m:
            local = m[0]
        procs = re.findall(r'"([^"]+)"', l)
        rows.append({
            "proto": proto,
            "local": local or (parts[4] if len(parts) > 4 else ""),
            "process": ", ".join(sorted(set(procs))) if procs else "",
        })
    return rows


def parse_software(lines):
    rows = []
    for l in lines:
        if "\t" in l:
            name, ver = l.split("\t", 1)
            rows.append((name.strip(), ver.strip()))
    return rows


def parse_users(user_lines, uid0_lines, sudoers_lines, pw_lines):
    nologin = ("nologin", "false", "sync", "halt", "shutdown")
    pw_state = {}
    for l in pw_lines:
        p = l.split()
        if len(p) >= 2:
            pw_state[p[0]] = p[1]  # P / L / NP
    sudo_users = set()
    for l in sudoers_lines:
        parts = l.split(":")
        if len(parts) >= 4 and parts[3]:
            sudo_users.update(x for x in parts[3].split(",") if x)
    uid0 = set(x.strip() for x in uid0_lines if x.strip())
    rows = []
    for l in user_lines:
        p = l.split(":")
        if len(p) < 4:
            continue
        name, uid, gid, shell = p[0], p[1], p[2], p[3]
        login = not any(shell.endswith(n) for n in nologin) and shell not in ("", "/bin/sync")
        st = {"P": "set", "L": "locked", "NP": "EMPTY"}.get(pw_state.get(name, ""), pw_state.get(name, ""))
        rows.append({
            "user": name, "uid": uid, "gid": gid, "shell": shell,
            "login_capable": "yes" if login else "no",
            "uid0": "YES" if name in uid0 else "",
            "sudo": "yes" if name in sudo_users else "",
            "password": st,
        })
    return rows, uid0


def parse_kv_colon(lines):
    d = {}
    for l in lines:
        if ":" in l:
            k, v = l.split(":", 1)
            d[k.strip().lower()] = v.strip()
    return d


def parse_ssh_effective(lines):
    d = {}
    for l in lines:
        p = l.split(None, 1)
        if len(p) == 2:
            d[p[0].lower()] = p[1].strip()
    return d


def parse_updates(lines):
    d = {"mgr": "", "pending": None, "security": None}
    for l in lines:
        if l.startswith("mgr:"):
            d["mgr"] = l.split(":", 1)[1]
        elif l.startswith("pending:"):
            try:
                d["pending"] = int(l.split(":", 1)[1])
            except ValueError:
                pass
        elif l.startswith("security:"):
            try:
                d["security"] = int(l.split(":", 1)[1])
            except ValueError:
                pass
    return d


def parse_sysctl(lines):
    d = {}
    for l in lines:
        if "=" in l:
            k, v = l.split("=", 1)
            d[k.strip()] = v.strip()
    return d


def parse_firewall(lines):
    active = False
    detail = []
    for l in lines:
        detail.append(l.strip())
        low = l.lower()
        if "ufw:" in low and "active" in low:
            active = True
        if "firewalld:" in low and "running" in low:
            active = True
        if low.startswith("nftables_lines:"):
            try:
                if int(l.split(":", 1)[1]) > 2:
                    active = True
            except ValueError:
                pass
        if low.startswith("iptables_rules:"):
            try:
                if int(l.split(":", 1)[1]) > 3:
                    active = True
            except ValueError:
                pass
    return active, "; ".join(x for x in detail if x)


# ----------------------------------------------------------------------------
# Build a normalised record + findings per host
# ----------------------------------------------------------------------------
def build_record(label, raw):
    sec = split_sections(raw)
    rec = {"host": label, "raw_sections": sec}

    rec["hostname"] = _first(sec.get("hostname", [])) or label
    rec["distro"] = parse_os_release(sec.get("os_release", []))
    kern = sec.get("kernel", [])
    rec["kernel"] = _first(kern)
    rec["arch"] = kern[1].strip() if len(kern) > 1 else ""
    rec["uptime"], rec["load"] = parse_uptime(sec.get("uptime", []))
    rec["ips"] = parse_ipaddr(sec.get("ipaddr", []))
    rec["disk"] = parse_disk(sec.get("disk", []))
    rec["listening"] = parse_listening(sec.get("listening", []))
    rec["services"] = [x.strip() for x in sec.get("services", []) if x.strip()]
    rec["software"] = parse_software(sec.get("software", []))
    rec["users"], rec["uid0"] = parse_users(
        sec.get("users", []), sec.get("uid0", []),
        sec.get("sudoers", []), sec.get("passwd_status", []),
    )
    rec["ssh"] = parse_ssh_effective(sec.get("ssh_effective", []))
    if not rec["ssh"]:  # fall back to raw file grep
        for l in sec.get("ssh_file", []):
            p = l.strip().split(None, 1)
            if len(p) == 2:
                rec["ssh"][p[0].lower()] = p[1]
    rec["fw_active"], rec["fw_detail"] = parse_firewall(sec.get("firewall", []))
    rec["mac"] = parse_kv_colon(sec.get("selinux", []))
    rec["updates"] = parse_updates(sec.get("updates", []))
    rec["reboot_required"] = _first(sec.get("reboot_required", [])) or "unknown"
    rec["sysctl"] = parse_sysctl(sec.get("sysctl", []))
    rec["pkgcount"] = _first(sec.get("pkgcount", []))
    rec["autoupdate"] = _first(sec.get("autoupdate", []))
    rec["logindefs"] = parse_logindefs(sec.get("logindefs", []))
    rec["insecure_pkgs"] = [x.strip() for x in sec.get("insecure_pkgs", []) if x.strip()]
    rec["file_perms"] = parse_file_perms(sec.get("file_perms", []))
    rec["packages"] = parse_software(sec.get("packages", []))  # name\tversion, same shape
    rec["timesync"] = parse_kv_equals(sec.get("timesync", []))

    # root disk %
    root_pct = 0
    for d in rec["disk"]:
        if d["mount"] == "/":
            try:
                root_pct = int(d["use_pct"].rstrip("%"))
            except ValueError:
                pass
    rec["root_pct"] = root_pct

    rec["findings"] = derive_findings(rec)
    rec["cis"] = derive_cis(rec)
    return rec


def parse_logindefs(lines):
    d = {}
    for l in lines:
        p = l.split()
        if len(p) >= 2:
            d[p[0].upper()] = p[1]
    return d


def parse_kv_equals(lines):
    d = {}
    for l in lines:
        if "=" in l:
            k, v = l.split("=", 1)
            d[k.strip()] = v.strip()
    return d


def parse_file_perms(lines):
    d = {}
    for l in lines:
        p = l.split()
        if len(p) >= 4:
            d[p[0]] = {"mode": p[1], "user": p[2], "group": p[3]}
    return d


def derive_cis(rec):
    """CIS-style benchmark checks. Each: (id, description, result, detail).
    result in {PASS, FAIL, WARN, N/A}. Derived from collected data only."""
    out = []

    def chk(cid, desc, result, detail=""):
        out.append((cid, desc, result, detail))

    def yn(v):
        return str(v).strip().lower()

    ssh = rec["ssh"]

    def sshval(k):
        return yn(ssh.get(k, ""))

    have_ssh = bool(ssh)
    # ---- SSH ----
    chk("5.2.1", "SSH: root login disabled",
        "PASS" if sshval("permitrootlogin") == "no" else ("N/A" if not have_ssh else "FAIL"),
        f"PermitRootLogin={ssh.get('permitrootlogin', '?')}")
    chk("5.2.2", "SSH: password authentication disabled",
        "PASS" if sshval("passwordauthentication") == "no" else ("N/A" if not have_ssh else "WARN"),
        f"PasswordAuthentication={ssh.get('passwordauthentication', '?')}")
    chk("5.2.3", "SSH: empty passwords not permitted",
        "PASS" if sshval("permitemptypasswords") == "no" else ("N/A" if not have_ssh else "FAIL"),
        f"PermitEmptyPasswords={ssh.get('permitemptypasswords', '?')}")
    chk("5.2.4", "SSH: MaxAuthTries <= 4",
        _num_cmp(ssh.get("maxauthtries"), lambda v: v <= 4),
        f"MaxAuthTries={ssh.get('maxauthtries', '?')}")
    chk("5.2.5", "SSH: X11 forwarding disabled",
        "PASS" if sshval("x11forwarding") == "no" else ("N/A" if not have_ssh else "WARN"),
        f"X11Forwarding={ssh.get('x11forwarding', '?')}")
    chk("5.2.6", "SSH: idle timeout configured (ClientAliveInterval 1-300)",
        _num_cmp(ssh.get("clientaliveinterval"), lambda v: 0 < v <= 300),
        f"ClientAliveInterval={ssh.get('clientaliveinterval', '?')}")
    chk("5.2.7", "SSH: LoginGraceTime <= 60",
        _num_cmp(ssh.get("logingracetime"), lambda v: v <= 60),
        f"LoginGraceTime={ssh.get('logingracetime', '?')}")

    # sshd_config file perms (0600 / root:root)
    fp = rec["file_perms"].get("/etc/ssh/sshd_config")
    if fp:
        ok = fp["mode"] in ("600", "644", "640") and fp["user"] == "root"
        chk("5.2.8", "SSH: sshd_config owned by root and not group/world-writable",
            "PASS" if (fp["user"] == "root" and int(fp["mode"], 8) & 0o022 == 0) else "FAIL",
            f"{fp['mode']} {fp['user']}:{fp['group']}")

    # ---- Filesystem / accounts perms ----
    for path, mode_ok, cid in [
        ("/etc/passwd", lambda m: int(m, 8) & 0o022 == 0, "6.1.2"),
        ("/etc/shadow", lambda m: int(m, 8) & 0o037 == 0, "6.1.3"),
        ("/etc/group", lambda m: int(m, 8) & 0o022 == 0, "6.1.4"),
        ("/etc/gshadow", lambda m: int(m, 8) & 0o037 == 0, "6.1.5"),
    ]:
        fp = rec["file_perms"].get(path)
        if fp:
            chk(cid, f"Permissions on {path} restricted",
                "PASS" if mode_ok(fp["mode"]) else "FAIL",
                f"{fp['mode']} {fp['user']}:{fp['group']}")

    chk("6.2.1", "Only one UID 0 account (root)",
        "PASS" if len(rec["uid0"]) <= 1 else "FAIL", ", ".join(sorted(rec["uid0"])))
    empties = [u["user"] for u in rec["users"]
               if u["password"] == "EMPTY" and u["login_capable"] == "yes"]
    chk("6.2.2", "No login accounts with empty password",
        "PASS" if not empties else "FAIL", ", ".join(empties) or "none")

    # ---- Password policy (login.defs) ----
    ld = rec["logindefs"]
    chk("5.4.1", "PASS_MAX_DAYS <= 365",
        _num_cmp(ld.get("PASS_MAX_DAYS"), lambda v: v <= 365, na_if_missing=True),
        f"PASS_MAX_DAYS={ld.get('PASS_MAX_DAYS', '?')}")
    chk("5.4.2", "PASS_MIN_DAYS >= 1",
        _num_cmp(ld.get("PASS_MIN_DAYS"), lambda v: v >= 1, na_if_missing=True),
        f"PASS_MIN_DAYS={ld.get('PASS_MIN_DAYS', '?')}")
    chk("5.4.3", "PASS_WARN_AGE >= 7",
        _num_cmp(ld.get("PASS_WARN_AGE"), lambda v: v >= 7, na_if_missing=True),
        f"PASS_WARN_AGE={ld.get('PASS_WARN_AGE', '?')}")

    # ---- Network / kernel hardening (sysctl) ----
    sc = rec["sysctl"]
    chk("1.5.1", "ASLR enabled (randomize_va_space=2)",
        "PASS" if sc.get("kernel.randomize_va_space") == "2" else "FAIL",
        f"randomize_va_space={sc.get('kernel.randomize_va_space', '?')}")
    chk("3.2.1", "IP forwarding disabled",
        "PASS" if sc.get("net.ipv4.ip_forward") == "0" else "WARN",
        f"ip_forward={sc.get('net.ipv4.ip_forward', '?')} (WARN=router/gateway ok)")
    chk("3.2.2", "ICMP redirects not accepted",
        "PASS" if sc.get("net.ipv4.conf.all.accept_redirects") == "0" else "FAIL",
        f"accept_redirects={sc.get('net.ipv4.conf.all.accept_redirects', '?')}")
    chk("3.2.3", "TCP SYN cookies enabled",
        "PASS" if sc.get("net.ipv4.tcp_syncookies") == "1" else "FAIL",
        f"tcp_syncookies={sc.get('net.ipv4.tcp_syncookies', '?')}")
    chk("3.2.4", "Reverse path filtering enabled",
        "PASS" if sc.get("net.ipv4.conf.all.rp_filter") in ("1", "2") else "WARN",
        f"rp_filter={sc.get('net.ipv4.conf.all.rp_filter', '?')}")

    # ---- MAC, firewall, time, patching ----
    selinux = yn(rec["mac"].get("selinux", ""))
    apparmor = yn(rec["mac"].get("apparmor", ""))
    mac_ok = selinux in ("enforcing", "permissive") or apparmor == "enabled"
    chk("1.6.1", "Mandatory Access Control active (SELinux/AppArmor)",
        "PASS" if mac_ok else "FAIL", f"selinux={selinux or 'n/a'} apparmor={apparmor or 'n/a'}")
    chk("3.5.1", "Host firewall active",
        "PASS" if rec["fw_active"] else "FAIL", rec["fw_detail"] or "none")
    ts = rec["timesync"]
    chk("2.2.1", "Time synchronization in use",
        "PASS" if ts.get("NTP") == "yes" or ts.get("NTPSynchronized") == "yes" else "WARN",
        f"NTP={ts.get('NTP', '?')} synced={ts.get('NTPSynchronized', '?')}")
    up = rec["updates"]
    chk("1.9", "No pending security updates",
        "PASS" if (up.get("security") in (0, None)) else "FAIL",
        f"security={up.get('security')} pending={up.get('pending')}")
    chk("1.8", "Automatic security updates configured",
        "PASS" if not rec["autoupdate"].startswith("auto-updates:none") else "WARN",
        rec["autoupdate"])
    chk("1.10", "No reboot pending",
        "PASS" if rec["reboot_required"] == "no" else ("N/A" if rec["reboot_required"] == "unknown" else "WARN"),
        rec["reboot_required"])

    # ---- Legacy insecure services ----
    chk("2.3", "No legacy insecure services installed (telnet/rsh/tftp/…)",
        "PASS" if not rec["insecure_pkgs"] else "FAIL",
        ", ".join(rec["insecure_pkgs"]) or "none")

    return out


def _num_cmp(val, predicate, na_if_missing=False):
    try:
        return "PASS" if predicate(int(str(val).strip())) else "FAIL"
    except (TypeError, ValueError):
        return "N/A" if na_if_missing else ("N/A" if val in (None, "") else "FAIL")


def derive_findings(rec):
    """Each finding: (severity, category, finding, detail, recommendation)."""
    f = []
    S_HIGH, S_MED, S_LOW, S_INFO = "High", "Medium", "Low", "Info"

    # Disk pressure
    for d in rec["disk"]:
        try:
            pct = int(d["use_pct"].rstrip("%"))
        except ValueError:
            continue
        if pct >= 90:
            f.append((S_HIGH, "Disk", f"Filesystem {d['mount']} {pct}% full",
                      f"{d['used']}/{d['size']} used", "Free space or expand volume; full disks cause outages"))
        elif pct >= 80:
            f.append((S_MED, "Disk", f"Filesystem {d['mount']} {pct}% full",
                      f"{d['used']}/{d['size']} used", "Plan capacity; approaching full"))

    # SSH hardening
    ssh = rec["ssh"]
    prl = ssh.get("permitrootlogin", "").lower()
    if prl in ("yes",):
        f.append((S_HIGH, "SSH", "PermitRootLogin is enabled", prl,
                  "Set 'PermitRootLogin no' and use sudo from a named account"))
    elif prl in ("prohibit-password", "without-password"):
        f.append((S_LOW, "SSH", "Root login allowed via key", prl,
                  "Prefer 'PermitRootLogin no' if not needed for automation"))
    if ssh.get("passwordauthentication", "").lower() == "yes":
        f.append((S_MED, "SSH", "Password authentication enabled", "passwordauthentication yes",
                  "Move to key-based auth and set 'PasswordAuthentication no'"))
    if ssh.get("permitemptypasswords", "").lower() == "yes":
        f.append((S_HIGH, "SSH", "Empty passwords permitted over SSH", "permitemptypasswords yes",
                  "Set 'PermitEmptyPasswords no' immediately"))
    if ssh.get("x11forwarding", "").lower() == "yes":
        f.append((S_LOW, "SSH", "X11 forwarding enabled", "x11forwarding yes",
                  "Disable unless required"))
    try:
        if int(ssh.get("maxauthtries", "6")) > 4:
            f.append((S_LOW, "SSH", "MaxAuthTries is high", ssh.get("maxauthtries"),
                      "Lower to 3-4 to slow brute forcing"))
    except ValueError:
        pass

    # Firewall
    if not rec["fw_active"]:
        f.append((S_MED, "Firewall", "No active host firewall detected", rec["fw_detail"] or "none",
                  "Enable ufw/firewalld/nftables with a default-deny inbound policy"))

    # MAC (SELinux/AppArmor)
    selinux = rec["mac"].get("selinux", "").lower()
    apparmor = rec["mac"].get("apparmor", "").lower()
    if selinux in ("disabled", "absent", "") and apparmor in ("disabled", "absent", ""):
        f.append((S_MED, "Kernel/MAC", "No mandatory access control active",
                  f"selinux={selinux or 'n/a'} apparmor={apparmor or 'n/a'}",
                  "Enable SELinux (enforcing) or AppArmor"))

    # Updates
    up = rec["updates"]
    if up.get("security"):
        f.append((S_HIGH, "Patching", f"{up['security']} pending security update(s)",
                  f"package manager: {up.get('mgr')}", "Apply security updates promptly"))
    elif up.get("pending"):
        f.append((S_MED, "Patching", f"{up['pending']} pending update(s)",
                  f"package manager: {up.get('mgr')}", "Schedule patching"))
    if rec["reboot_required"] == "yes":
        f.append((S_MED, "Patching", "Reboot required", "kernel/libs updated",
                  "Schedule a maintenance reboot"))
    if rec["autoupdate"].startswith("auto-updates:none"):
        f.append((S_LOW, "Patching", "No automatic security updates configured", rec["autoupdate"],
                  "Consider unattended-upgrades / dnf-automatic for security patches"))

    # Accounts
    if len(rec["uid0"]) > 1:
        f.append((S_HIGH, "Accounts", "Multiple UID 0 accounts", ", ".join(sorted(rec["uid0"])),
                  "Only 'root' should have UID 0; investigate the others"))
    empties = [u["user"] for u in rec["users"] if u["password"] == "EMPTY" and u["login_capable"] == "yes"]
    if empties:
        f.append((S_HIGH, "Accounts", "Login account(s) with empty password", ", ".join(empties),
                  "Lock or set passwords on these accounts"))

    # Sysctl hardening
    sc = rec["sysctl"]
    if sc.get("kernel.randomize_va_space") not in (None, "2", ""):
        f.append((S_MED, "Kernel", "ASLR not fully enabled",
                  f"kernel.randomize_va_space={sc.get('kernel.randomize_va_space')}",
                  "Set kernel.randomize_va_space=2"))
    if sc.get("net.ipv4.tcp_syncookies") == "0":
        f.append((S_LOW, "Kernel", "TCP SYN cookies disabled", "net.ipv4.tcp_syncookies=0",
                  "Enable to mitigate SYN floods"))
    if sc.get("net.ipv4.ip_forward") == "1":
        f.append((S_INFO, "Kernel", "IP forwarding enabled", "net.ipv4.ip_forward=1",
                  "Expected on routers/gateways; disable otherwise"))

    return f


# ----------------------------------------------------------------------------
# Excel writer
# ----------------------------------------------------------------------------
def make_workbook(records, errors, args):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT = "Arial"
    hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    base_font = Font(name=FONT, size=10)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sev_fill = {
        "High": PatternFill("solid", fgColor=RED),
        "Medium": PatternFill("solid", fgColor=ORANGE),
        "Low": PatternFill("solid", fgColor=YELLOW),
        "Info": PatternFill("solid", fgColor=GREY),
    }
    sev_font = {
        "High": Font(name=FONT, bold=True, color="FFFFFF", size=10),
        "Medium": Font(name=FONT, bold=True, color="FFFFFF", size=10),
        "Low": Font(name=FONT, size=10),
        "Info": Font(name=FONT, size=10),
    }

    wb = Workbook()

    def add_sheet(title, headers, rows, widths, cell_style=None, freeze="A2"):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            cell.border = border
        for r, row in enumerate(rows, start=2):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                _guard_formula(cell)
                cell.font = base_font
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                cell.border = border
                if cell_style:
                    cell_style(ws, r, c, headers, row)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = freeze
        if rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        ws.sheet_view.showGridLines = False
        return ws

    # ---- Summary -----------------------------------------------------------
    sum_hdr = ["Host", "Hostname", "Primary IP", "Distro", "Kernel", "Arch",
               "Uptime", "Load (1/5/15m)", "Root Disk %", "Listening Ports",
               "Firewall", "SELinux/AppArmor", "Root SSH Login",
               "Password Auth", "Pending Updates", "Security Updates",
               "Reboot Req.", "Packages", "CIS Fail/Warn", "Findings (H/M/L)"]
    sum_rows = []
    for r in records:
        sev_counts = {"High": 0, "Medium": 0, "Low": 0, "Info": 0}
        for s, *_ in r["findings"]:
            sev_counts[s] = sev_counts.get(s, 0) + 1
        cis_fail = sum(1 for c in r["cis"] if c[2] == "FAIL")
        cis_warn = sum(1 for c in r["cis"] if c[2] == "WARN")
        mac = f"{r['mac'].get('selinux', 'n/a')}/{r['mac'].get('apparmor', 'n/a')}"
        sum_rows.append([
            r["host"], r["hostname"], (r["ips"][0] if r["ips"] else ""),
            r["distro"], r["kernel"], r["arch"], r["uptime"], r["load"],
            r["root_pct"], len(r["listening"]),
            "active" if r["fw_active"] else "INACTIVE", mac,
            r["ssh"].get("permitrootlogin", "?"),
            r["ssh"].get("passwordauthentication", "?"),
            (r["updates"].get("pending") if r["updates"].get("pending") is not None else ""),
            (r["updates"].get("security") if r["updates"].get("security") is not None else ""),
            r["reboot_required"], r["pkgcount"],
            f"{cis_fail}/{cis_warn}",
            f"{sev_counts['High']}/{sev_counts['Medium']}/{sev_counts['Low']}",
        ])

    def summary_style(ws, row, col, headers, rowvals):
        h = headers[col - 1]
        cell = ws.cell(row=row, column=col)
        if h == "Root Disk %":
            try:
                v = int(cell.value)
                if v >= 90:
                    cell.fill = sev_fill["High"]; cell.font = sev_font["High"]
                elif v >= 80:
                    cell.fill = sev_fill["Medium"]; cell.font = sev_font["Medium"]
            except (TypeError, ValueError):
                pass
        elif h == "Firewall" and cell.value == "INACTIVE":
            cell.fill = sev_fill["Medium"]; cell.font = sev_font["Medium"]
        elif h == "Root SSH Login" and str(cell.value).lower() == "yes":
            cell.fill = sev_fill["High"]; cell.font = sev_font["High"]
        elif h == "Password Auth" and str(cell.value).lower() == "yes":
            cell.fill = sev_fill["Medium"]; cell.font = sev_font["Medium"]
        elif h == "Security Updates":
            try:
                if int(cell.value) > 0:
                    cell.fill = sev_fill["High"]; cell.font = sev_font["High"]
            except (TypeError, ValueError):
                pass
        elif h == "Reboot Req." and cell.value == "yes":
            cell.fill = sev_fill["Medium"]; cell.font = sev_font["Medium"]
        elif h == "CIS Fail/Warn":
            try:
                fail = int(str(cell.value).split("/")[0])
                if fail > 0:
                    cell.fill = sev_fill["High"]; cell.font = sev_font["High"]
            except (TypeError, ValueError):
                pass

    add_sheet("Summary", sum_hdr, sum_rows,
              [22, 24, 15, 30, 20, 8, 16, 14, 11, 14, 12, 18, 15, 14, 14, 15, 11, 10, 13, 14],
              cell_style=summary_style)

    # ---- Findings ----------------------------------------------------------
    sev_order = {"High": 0, "Medium": 1, "Low": 2, "Info": 3}
    find_rows = []
    for r in records:
        for s, cat, finding, detail, rec_txt in r["findings"]:
            find_rows.append([r["host"], s, cat, finding, detail, rec_txt])
    find_rows.sort(key=lambda x: (x[0], sev_order.get(x[1], 9)))

    def findings_style(ws, row, col, headers, rowvals):
        if headers[col - 1] == "Severity":
            cell = ws.cell(row=row, column=col)
            sv = cell.value
            if sv in sev_fill:
                cell.fill = sev_fill[sv]; cell.font = sev_font[sv]

    add_sheet("Findings",
              ["Host", "Severity", "Category", "Finding", "Detail", "Recommendation"],
              find_rows, [22, 10, 14, 40, 40, 55], cell_style=findings_style)

    # ---- CIS Checks --------------------------------------------------------
    cis_fill = {
        "PASS": PatternFill("solid", fgColor=GREEN),
        "FAIL": PatternFill("solid", fgColor=RED),
        "WARN": PatternFill("solid", fgColor=ORANGE),
        "N/A": PatternFill("solid", fgColor=GREY),
    }
    cis_txtfont = {
        "PASS": Font(name=FONT, bold=True, color="FFFFFF", size=10),
        "FAIL": Font(name=FONT, bold=True, color="FFFFFF", size=10),
        "WARN": Font(name=FONT, bold=True, color="FFFFFF", size=10),
        "N/A": Font(name=FONT, size=10),
    }
    res_order = {"FAIL": 0, "WARN": 1, "N/A": 2, "PASS": 3}
    cis_rows = []
    for r in records:
        for cid, desc, result, detail in r["cis"]:
            cis_rows.append([r["host"], cid, desc, result, detail])
    cis_rows.sort(key=lambda x: (x[0], res_order.get(x[3], 9), x[1]))

    def cis_style(ws, row, col, headers, rowvals):
        if headers[col - 1] == "Result":
            cell = ws.cell(row=row, column=col)
            if cell.value in cis_fill:
                cell.fill = cis_fill[cell.value]
                cell.font = cis_txtfont[cell.value]
                cell.alignment = Alignment(horizontal="center", vertical="top")

    add_sheet("CIS Checks",
              ["Host", "Check ID", "Description", "Result", "Detail"],
              cis_rows, [22, 10, 52, 10, 46], cell_style=cis_style)

    # ---- Disk --------------------------------------------------------------
    disk_rows = []
    for r in records:
        for d in r["disk"]:
            disk_rows.append([r["host"], d["filesystem"], d["type"], d["size"],
                              d["used"], d["avail"], d["use_pct"], d["mount"]])

    def disk_style(ws, row, col, headers, rowvals):
        if headers[col - 1] == "Use%":
            cell = ws.cell(row=row, column=col)
            try:
                v = int(str(cell.value).rstrip("%"))
                if v >= 90:
                    cell.fill = sev_fill["High"]; cell.font = sev_font["High"]
                elif v >= 80:
                    cell.fill = sev_fill["Medium"]; cell.font = sev_font["Medium"]
            except (TypeError, ValueError):
                pass

    add_sheet("Disk",
              ["Host", "Filesystem", "Type", "Size", "Used", "Avail", "Use%", "Mount"],
              disk_rows, [22, 26, 10, 10, 10, 10, 8, 26], cell_style=disk_style)

    # ---- Listening ports ---------------------------------------------------
    port_rows = []
    for r in records:
        for p in r["listening"]:
            port_rows.append([r["host"], p["proto"], p["local"], p["process"]])
    add_sheet("Listening Ports",
              ["Host", "Proto", "Local Address:Port", "Process"],
              port_rows, [22, 8, 30, 40])

    # ---- Services ----------------------------------------------------------
    svc_rows = [[r["host"], s] for r in records for s in r["services"]]
    add_sheet("Services", ["Host", "Running Service"], svc_rows, [22, 50])

    # ---- Software versions -------------------------------------------------
    sw_rows = [[r["host"], n, v] for r in records for n, v in r["software"]]
    add_sheet("Software Versions", ["Host", "Software", "Version"],
              sw_rows, [22, 18, 70])

    # ---- Installed packages (full inventory) -------------------------------
    if getattr(args, "packages", True):
        pkg_rows = [[r["host"], n, v] for r in records for n, v in r["packages"]]
        if pkg_rows:
            add_sheet("Installed Packages", ["Host", "Package", "Version"],
                      pkg_rows, [22, 34, 34])

    # ---- Users & Auth ------------------------------------------------------
    user_rows = []
    for r in records:
        for u in r["users"]:
            if u["login_capable"] == "yes" or u["uid0"] or u["sudo"]:
                user_rows.append([r["host"], u["user"], u["uid"], u["gid"],
                                  u["shell"], u["login_capable"], u["uid0"],
                                  u["sudo"], u["password"]])

    def user_style(ws, row, col, headers, rowvals):
        cell = ws.cell(row=row, column=col)
        h = headers[col - 1]
        if h == "UID0" and cell.value == "YES":
            cell.fill = sev_fill["High"]; cell.font = sev_font["High"]
        elif h == "Password" and cell.value == "EMPTY":
            cell.fill = sev_fill["High"]; cell.font = sev_font["High"]

    add_sheet("Users & Auth",
              ["Host", "User", "UID", "GID", "Shell", "Login", "UID0", "Sudo", "Password"],
              user_rows, [22, 18, 8, 8, 22, 8, 8, 8, 12], cell_style=user_style)

    # ---- SSH config --------------------------------------------------------
    ssh_keys = ["permitrootlogin", "passwordauthentication", "pubkeyauthentication",
                "permitemptypasswords", "x11forwarding", "maxauthtries",
                "clientaliveinterval", "logingracetime", "allowtcpforwarding"]
    ssh_hdr = ["Host"] + [k for k in ssh_keys]
    ssh_rows = [[r["host"]] + [r["ssh"].get(k, "") for k in ssh_keys] for r in records]
    add_sheet("SSH Config", ssh_hdr, ssh_rows,
              [22] + [20] * len(ssh_keys))

    # ---- Per-host detail tabs ---------------------------------------------
    if getattr(args, "host_tabs", True):
        used_names = set()

        def safe_sheet_name(name):
            clean = re.sub(r"[\[\]\:\*\?\/\\]", "-", name)[:31]
            base, i = clean, 1
            while clean.lower() in used_names or not clean:
                suffix = f"~{i}"
                clean = (base[:31 - len(suffix)] + suffix)
                i += 1
            used_names.add(clean.lower())
            return clean

        sec_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        sec_fill = PatternFill("solid", fgColor="1F3864")
        key_font = Font(name=FONT, bold=True, size=10)
        val_font = Font(name=FONT, size=10)

        for r in records:
            ws = wb.create_sheet(safe_sheet_name(r["hostname"] or r["host"]))
            ws.sheet_view.showGridLines = False
            ws.column_dimensions["A"].width = 26
            ws.column_dimensions["B"].width = 90
            row = [1]

            def title(txt):
                c = ws.cell(row=row[0], column=1, value=txt)
                _guard_formula(c)
                c.font = Font(name=FONT, bold=True, size=14, color="1F3864")
                row[0] += 1

            def section(txt):
                row[0] += 1
                for col in (1, 2):
                    cc = ws.cell(row=row[0], column=col, value=txt if col == 1 else "")
                    cc.font = sec_font
                    cc.fill = sec_fill
                row[0] += 1

            def kv(k, v, fill=None, fillfont=None):
                a = ws.cell(row=row[0], column=1, value=k)
                b = ws.cell(row=row[0], column=2, value=v)
                _guard_formula(a)
                _guard_formula(b)
                a.font = key_font
                b.font = fillfont or val_font
                b.alignment = Alignment(wrap_text=True, vertical="top")
                if fill:
                    b.fill = fill
                row[0] += 1

            title(f"{r['hostname']}  ({r['host']})")
            section("System")
            kv("Distro", r["distro"])
            kv("Kernel / Arch", f"{r['kernel']}  {r['arch']}")
            kv("Uptime", r["uptime"])
            kv("Load (1/5/15m)", r["load"])
            kv("Installed packages", r["pkgcount"])
            kv("Reboot required", r["reboot_required"],
               fill=(sev_fill["Medium"] if r["reboot_required"] == "yes" else None),
               fillfont=(sev_font["Medium"] if r["reboot_required"] == "yes" else None))

            section("Network")
            kv("IP addresses", ", ".join(r["ips"]) or "—")
            for p in r["listening"]:
                kv(f"Listening {p['proto']}", f"{p['local']}  ->  {p['process'] or '?'}")

            section("Disk")
            for d in r["disk"]:
                try:
                    pct = int(d["use_pct"].rstrip("%"))
                except ValueError:
                    pct = 0
                fill = sev_fill["High"] if pct >= 90 else (sev_fill["Medium"] if pct >= 80 else None)
                ff = sev_font["High"] if pct >= 90 else (sev_font["Medium"] if pct >= 80 else None)
                kv(f"{d['mount']} ({d['type']})",
                   f"{d['used']}/{d['size']} used ({d['use_pct']}), {d['avail']} free",
                   fill=fill, fillfont=ff)

            section("Security posture")
            fw = "active" if r["fw_active"] else "INACTIVE"
            kv("Firewall", fw,
               fill=(None if r["fw_active"] else sev_fill["Medium"]),
               fillfont=(None if r["fw_active"] else sev_font["Medium"]))
            kv("SELinux / AppArmor",
               f"{r['mac'].get('selinux', 'n/a')} / {r['mac'].get('apparmor', 'n/a')}")
            kv("SSH PermitRootLogin", r["ssh"].get("permitrootlogin", "?"))
            kv("SSH PasswordAuth", r["ssh"].get("passwordauthentication", "?"))
            kv("Pending / security updates",
               f"{r['updates'].get('pending')} / {r['updates'].get('security')}")
            cis_fail_n = sum(1 for c in r["cis"] if c[2] == "FAIL")
            cis_warn_n = sum(1 for c in r["cis"] if c[2] == "WARN")
            cis_pass_n = sum(1 for c in r["cis"] if c[2] == "PASS")
            kv("CIS pass / fail / warn", f"{cis_pass_n} / {cis_fail_n} / {cis_warn_n}",
               fill=(sev_fill["High"] if cis_fail_n else None),
               fillfont=(sev_font["High"] if cis_fail_n else None))

            section("Accounts (login-capable / privileged)")
            for u in r["users"]:
                if u["login_capable"] == "yes" or u["uid0"] or u["sudo"]:
                    tags = []
                    if u["uid0"]:
                        tags.append("UID0")
                    if u["sudo"]:
                        tags.append("sudo")
                    tags.append(f"pw:{u['password']}")
                    kv(u["user"], f"uid={u['uid']} shell={u['shell']}  [{', '.join(tags)}]")

            section("Top findings")
            hi = [f for f in r["findings"] if f[0] in ("High", "Medium")]
            if not hi:
                kv("—", "No high/medium findings")
            for s, cat, finding, detail, rec_txt in hi:
                kv(s, f"[{cat}] {finding} — {rec_txt}",
                   fill=sev_fill.get(s), fillfont=sev_font.get(s))

            ws.freeze_panes = "A2"

    # ---- Errors ------------------------------------------------------------
    if errors:
        add_sheet("Errors", ["Host", "Error"], errors, [28, 90])

    # ---- Cover / metadata --------------------------------------------------
    cover = wb.active
    cover.title = "About"
    cover.sheet_view.showGridLines = False
    meta = [
        ("Linux Fleet Audit Report", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Script build", BUILD),
        ("Hosts audited", str(len(records))),
        ("Hosts failed", str(len(errors))),
        ("Total findings", str(sum(len(r["findings"]) for r in records))),
        ("Total CIS FAILs", str(sum(1 for r in records for c in r["cis"] if c[2] == "FAIL"))),
        ("Escalation", args.escalate),
        ("Deep scan (SUID/world-writable)", "yes" if args.deep else "no"),
        ("Full package inventory", "yes" if getattr(args, "packages", True) else "no"),
        ("Per-host detail tabs", "yes" if getattr(args, "host_tabs", True) else "no"),
        ("", ""),
        ("Sheet guide", ""),
        ("Summary", "One row per host with key facts and colour-coded risk cells"),
        ("Findings", "Prioritised hardening findings with severity and recommendations"),
        ("CIS Checks", "CIS-style PASS/FAIL/WARN benchmark checks per host"),
        ("Disk / Listening Ports / Services / Software Versions", "Raw inventory detail"),
        ("Installed Packages", "Full installed-package inventory per host"),
        ("Users & Auth / SSH Config", "Account and SSH hardening detail"),
        ("<hostname> tabs", "One detail tab per host (system, network, disk, posture)"),
        ("Errors", "Hosts that could not be reached or collected"),
        ("", ""),
        ("Note", "Update counts reflect the package cache as-is (no repo refresh was run)."),
        ("Note", "CIS checks are heuristic, benchmark-style; not a certified CIS-CAT scan."),
    ]
    cover["A1"].font = Font(name=FONT, bold=True, size=16, color="1F3864")
    for i, (k, v) in enumerate(meta, start=1):
        a = cover.cell(row=i, column=1, value=k)
        b = cover.cell(row=i, column=2, value=v)
        if i == 1:
            continue
        a.font = Font(name=FONT, bold=True, size=10)
        b.font = Font(name=FONT, size=10)
    cover.column_dimensions["A"].width = 44
    cover.column_dimensions["B"].width = 70
    wb.move_sheet("About", -(len(wb.sheetnames) - 1))

    # Final safety sweep: no cell may be saved as a formula. Any value openpyxl
    # typed as a formula (string starting with '=') is forced back to text. This
    # guarantees Excel never emits a "Removed Records: Formula" repair warning and
    # neutralises spreadsheet formula injection from untrusted host data.
    swept = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    cell.data_type = "s"
                    swept += 1
    if swept:
        print(f"  (formula-safety sweep neutralised {swept} cell(s))", file=sys.stderr)

    wb.save(args.output)


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main():
    print_banner()
    ap = argparse.ArgumentParser(
        description="SSH Linux fleet inventory & hardening baseline -> Excel.")
    ap.add_argument("-H", "--hosts", required=True, help="text file, one host per line")
    ap.add_argument("-o", "--output", default="linux_audit.xlsx", help="output .xlsx path")
    ap.add_argument("--active-out", default="active_hosts.txt", metavar="PATH",
                    help="write reachable hosts to this file (default: active_hosts.txt; "
                         "pass '' to disable)")
    ap.add_argument("-u", "--user", default=None, help="default SSH user (overridable per-line)")
    ap.add_argument("-p", "--port", default=None, help="default SSH port")
    ap.add_argument("-i", "--identity", default=None, help="SSH private key file")
    ap.add_argument("--escalate", choices=["none", "sudo"], default="sudo",
                    help="privilege escalation method (default: sudo)")
    ap.add_argument("--ask-sudo-pass", action="store_true",
                    help="prompt once for a sudo password (fed via sudo -S)")
    ap.add_argument("--sudo-pass-same-as-ssh", action="store_true",
                    help="reuse the SSH login password for sudo (no second prompt)")
    ap.add_argument("--ask-ssh-pass", action="store_true",
                    help="use password SSH login via sshpass (prompt once)")
    ap.add_argument("--ssh-pass-env", default=None, metavar="VAR",
                    help="read the SSH login password from this env var instead of prompting")
    ap.add_argument("--ssh-opt", action="append",
                    help="extra ssh -o option, repeatable (e.g. --ssh-opt ProxyJump=bastion)")
    ap.add_argument("--host-key-checking", default="accept-new",
                    help="StrictHostKeyChecking value (default: accept-new)")
    ap.add_argument("--connect-timeout", type=int, default=10, help="SSH connect timeout (s)")
    ap.add_argument("--cmd-timeout", type=int, default=120, help="per-host command timeout (s)")
    ap.add_argument("--workers", type=int, default=8, help="parallel hosts (default: 8)")
    ap.add_argument("--deep", action="store_true",
                    help="also scan SUID/SGID and world-writable files (slower)")
    ap.add_argument("--packages", action=argparse.BooleanOptionalAction, default=True,
                    help="collect full installed-package inventory (default: on)")
    ap.add_argument("--host-tabs", action=argparse.BooleanOptionalAction, default=True,
                    help="add one detail tab per host (default: on)")
    args = ap.parse_args()

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        sys.exit("Missing dependency: pip install openpyxl")

    hosts = parse_hosts(args.hosts)
    if not hosts:
        sys.exit("No hosts found in " + args.hosts)

    # SSH password login (via sshpass)
    ssh_pass = None
    if args.ssh_pass_env:
        ssh_pass = os.environ.get(args.ssh_pass_env)
        if ssh_pass is None:
            sys.exit(f"Env var {args.ssh_pass_env} is not set")
    elif args.ask_ssh_pass:
        ssh_pass = getpass.getpass("SSH login password (reused for all hosts): ")
    if ssh_pass is not None:
        import shutil
        if not shutil.which("sshpass"):
            sys.exit("Password SSH login needs 'sshpass'. Install it "
                     "(apt install sshpass / dnf install sshpass) or use key-based auth.")

    sudo_pass = None
    if args.escalate == "sudo":
        if args.sudo_pass_same_as_ssh:
            if ssh_pass is None:
                sys.exit("--sudo-pass-same-as-ssh requires --ask-ssh-pass or --ssh-pass-env")
            sudo_pass = ssh_pass
        elif args.ask_sudo_pass:
            sudo_pass = getpass.getpass("sudo password (reused for all hosts): ")

    script = (("export DEEP=1\n" if args.deep else "")
              + ("export PKGS=1\n" if args.packages else "")
              + COLLECTOR).encode()

    print(f"Auditing {len(hosts)} host(s) with {args.workers} worker(s)...  [build {BUILD}]",
          file=sys.stderr)

    records, errors = [], []

    def work(h):
        label = (f"{h['user']}@" if h["user"] else "") + h["target"] + \
                (f":{h['port']}" if h["port"] else "")
        raw, err = collect_host(h, args, script, sudo_pass, ssh_pass)
        return label, raw, err

    with cf.ThreadPoolExecutor(max_workers=args.workers) as ex:
        for label, raw, err in ex.map(work, hosts):
            if err and "@@SECTION" not in raw:
                print(f"  [FAIL] {label}: {err}", file=sys.stderr)
                errors.append([label, err])
            else:
                print(f"  [ OK ] {label}", file=sys.stderr)
                records.append(build_record(label, raw))

    records.sort(key=lambda r: r["host"])
    if not records:
        print("No hosts succeeded; writing errors-only report.", file=sys.stderr)

    make_workbook(records, errors, args)
    print(f"\nWrote {args.output}  ({len(records)} ok, {len(errors)} failed)",
          file=sys.stderr)

    if args.active_out:
        write_active_file(records, errors, args.active_out)
        print(f"Wrote {args.active_out}  ({len(records)} active host(s))",
              file=sys.stderr)


def write_active_file(records, errors, path):
    """Write reachable hosts as a reusable hosts file (comments are ignored on
    re-read, so each line stays valid input for a later run)."""
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(path, "w") as fh:
        fh.write(f"# Active hosts as of {stamp} — "
                 f"{len(records)} reachable, {len(errors)} unreachable\n")
        fh.write("# Host token is first; trailing '# ...' is just annotation.\n")
        for r in records:
            ip = r["ips"][0] if r["ips"] else "?"
            note = f"{ip}  {r['distro']}".strip()
            fh.write(f"{r['host']:<32}# {note}\n")



if __name__ == "__main__":
    main()
