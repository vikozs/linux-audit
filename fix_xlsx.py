#!/usr/bin/env python3
"""Repair an existing .xlsx that Excel flags with 'Removed Records: Formula'.

Converts every cell openpyxl typed as a formula (any value starting with '=')
back to plain text, preserving all data and formatting. Use on reports made by
an older linux_audit.py build.

    python3 fix_xlsx.py linux_audit.xlsx            # writes linux_audit_fixed.xlsx
    python3 fix_xlsx.py linux_audit.xlsx clean.xlsx # explicit output name
"""
import sys
from openpyxl import load_workbook

if len(sys.argv) < 2:
    sys.exit("usage: python3 fix_xlsx.py <input.xlsx> [output.xlsx]")
inp = sys.argv[1]
out = sys.argv[2] if len(sys.argv) > 2 else inp.rsplit(".", 1)[0] + "_fixed.xlsx"

wb = load_workbook(inp)
n = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                cell.data_type = "s"   # keep the text, drop the formula typing
                n += 1
wb.save(out)
print(f"Neutralised {n} formula cell(s). Wrote {out}")
